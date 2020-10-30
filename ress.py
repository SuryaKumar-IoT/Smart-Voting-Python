from openpyxl import load_workbook
# Give the location of the file 
loc = ("result.xlsx") 

def updateres(a,b,c):
    wb = load_workbook(loc)
    ws = wb.get_sheet_by_name("Sheet1")
    ws.cell(row=a,column=b).value = c
    wb.save('result.xlsx')

#updateres(3,1,45)
